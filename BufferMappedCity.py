import requests
from bs4 import BeautifulSoup, Comment, NavigableString
import sys, codecs, json
import openpyxl

# global variables
distanceThreshold = 200000 # unit is in meters here
stationList = "c:/github/airQuality/DownloadedData/Station_List.xlsx"
districtList = "c:/github/airQuality/DistrictWiseData.xlsx"
# API key below can be called just from my IP - its super-restricted - can change settigs if needed
requestUrl = "https://maps.googleapis.com/maps/api/distancematrix/json?origins={}&destinations={}&key=AIzaSyCAanMVI-hfmXrv_-xKh3KHI-x-OqdGoJM"

# Buffer all the state-city pair that we have
StationWB = openpyxl.load_workbook(stationList)
stateCityPairSheet = StationWB["StateCityPair"] # all city state pair with active stations
max_row = stateCityPairSheet.max_row
stateCityPair = {}
stateToCityMap = {}

for i in range(2, max_row + 1):    
    state = str(stateCityPairSheet.cell(row = i, column = 1).value).lower() 
    city = str(stateCityPairSheet.cell(row = i, column = 2).value).lower()
    cell_obj = state + city
    stateCityPair[cell_obj] = True
    if not state in stateToCityMap:
        stateToCityMap[state] = list()
    stateToCityMap[state].append(city)

# Fill in cities for available data
districtWB = openpyxl.load_workbook(districtList)
districtSheet = districtWB["List of Districts"] # Get list of all districts
max_row = districtSheet.max_row
count = 0

for i in range(2, max_row + 1): 
    state = str(districtSheet.cell(row = i, column = 1).value).lower()
    city = str(districtSheet.cell(row = i, column = 3).value).lower()
    cell_obj = state + city
    if cell_obj in stateCityPair:
        districtSheet.cell(row = i, column = 4).value = districtSheet.cell(row = i, column = 3).value
    else:
        # Fill in proxy city for missing cities 
        if state in stateToCityMap:
            # find distance here
            nearestDistance = -1
            curDistance = 0
            nearestCity = ""
            for cityData in stateToCityMap[state]:
                address1 = ",".join([city, state])
                address2 = ",".join([cityData, state])
                formattedRequestUrl = requestUrl.format(address1, address2 )
                r = requests.get(formattedRequestUrl)
                site_json=json.loads(r.text)
                if site_json['rows']:
                    rows = site_json['rows']
                    if rows[0]['elements']:
                        elements = rows[0]['elements']
                        if elements[0]["distance"]:
                            distance = elements[0]["distance"]
                            curDistance = int(distance["value"])
                            if nearestDistance == -1 or curDistance < nearestDistance:
                                nearestDistance = curDistance
                                nearestCity = cityData
            # Only update to nearest city if this meets the threshold limit
            if(nearestDistance <= distanceThreshold): 
                districtSheet.cell(row = i, column = 4).value = nearestCity
        # else: Nothing in case the data for a state does not exist

districtWB.save(districtList)
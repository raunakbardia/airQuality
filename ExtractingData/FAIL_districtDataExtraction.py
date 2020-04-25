import requests
from bs4 import BeautifulSoup, Comment, NavigableString
import sys, codecs, json
import openpyxl

# Load the workbook
path = "DistrictWiseData.xlsx"
wb = openpyxl.load_workbook(path)
stateSheet = wb["List of States"]   # separate sheet with stats about each state
districtSheet = wb["List of Districts"] # separate sheet with stats about each district
nStates = len(stateSheet['A'])      # Number of rows in the state sheet to run nStates iterations

# For loop to extract district data for each state from specific URLs
for rowNum in range(nStates-2):
    currentState = rowNum + 2
    currentDistrict = rowNum + 2
    state = stateSheet.cell(row = currentState, column = 2) # Cell value is stored in cell format
    stateName = state.value.replace(" ","_")    # Spaces replaced by underscore to be used in hyperlink
    print("Reading data for ", state.value)
    s = BeautifulSoup(requests.get('https://en.wikipedia.org/wiki/List_of_districts_of_{0}'.format(stateName)).text, 'html.parser')
    myTable = s.findAll('table',{'class':'sortable'})   # Lookup table in the hyperlink
    
    if(len(myTable) == 0):
        print("District table not found for ", state.value)
        continue

    colNumber = 0
    for header in myTable[0].findAll('th'):     # Lookup header tags in the table
        if('District' in header.text):          # Columns where specific values are stored in a table may differ
            districtCol = colNumber
        else:
            if('Population' in header.text):
                populationCol = colNumber
            else:
                if('Density' in header.text):
                    densityCol = colNumber
                else:
                    if('Area' in header.text):
                        areaCol = colNumber
        colNumber = colNumber + 1

    for rows in myTable[0].findAll('tr'):       # Read each row and extract specific columns
        colNumber = 0                           # Counter to keep track on iterated number of columns for each row
        for cols in rows.findAll('td'):
            if(colNumber == districtCol):
                c1 = districtSheet.cell(row = currentDistrict, column = 1)
                c1.value = state.value
                c2 = districtSheet.cell(row = currentDistrict, column = 2)
                c2.value = cols.text
            else:
                if(colNumber == populationCol):
                    c3 = districtSheet.cell(row=currentDistrict, column = 3)
                    c3.value = cols.text
                else:
                    if(colNumber == areaCol):
                        c4 = districtSheet.cell(row=currentDistrict, column = 4)
                        c4.value = cols.text
                    else:
                        if(colNumber == densityCol):
                            c5 = districtSheet.cell(row=currentDistrict, column = 5)
                            c5.value = cols.text
            colNumber = colNumber + 1
        if(colNumber != 0):
            currentDistrict = currentDistrict + 1
wb.save("DistrictWiseData.xlsx")
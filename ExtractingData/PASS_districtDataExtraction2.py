import requests
from bs4 import BeautifulSoup, Comment, NavigableString
import sys, codecs, json
import openpyxl

# Load the workbook
path = "DistrictWiseData.xlsx"
wb = openpyxl.load_workbook(path)
stateSheet = wb["List of States"]   # separate sheet with stats about each state
districtSheet = wb["List of Districts"] # separate sheet with stats about each district
nStates = len(stateSheet['A'])      # Number of rows in the state sheet to run nStates iterations

r = requests.get('https://en.wikipedia.org/wiki/List_of_districts_in_India')
s = BeautifulSoup(r.text, 'html.parser')
myTable = s.findAll('table',{'class':'wikitable sortable'})
nTables = len(myTable)

# Length of myTable is 37: 1 Initial State Table + 36 District Tables
# Goal: To read the 36 tables
# Format of these 36 tables is consistent
# Required data in 3rd, 5th, 6th, and 7th columns
# Col 3 - District
# Col 5 - Population
# Col 6 - Area
# Col 7 - Density
nDistricts = 0
excelRow = 2
for x in range(1,nTables):
    activeTable = myTable[x]
    stateName = stateSheet.cell(row = x + 1, column = 2).value
    print("Reading district data for ", stateName)
    for row in activeTable.find_all('tr'):
        colNumber = 0
        for col in row.find_all('td'):
            activeCol = 0
            if(colNumber == 2):
                activeCol = 2
            else:
                if(colNumber == 4):
                    activeCol = 3
                else:
                    if(colNumber == 5):
                        activeCol = 4
                    else:
                        if(colNumber == 6):
                            activeCol = 5
            if(activeCol != 0):
                val = col.text.replace(" ","")
                start = val.find("[")
                end = val.find("]")
                if start != -1 and end != -1:
                    result = val[start:end+1]
                    val = val.replace(result,"")
                districtSheet.cell(row = excelRow, column = activeCol).value = val
            colNumber = colNumber + 1
        if(colNumber > 0):
            districtSheet.cell(row = excelRow, column = 1).value = stateName
            excelRow = excelRow + 1
wb.save(path)
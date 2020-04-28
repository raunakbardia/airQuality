from selenium import webdriver
import time
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select, WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
import requests
from bs4 import BeautifulSoup, Comment, NavigableString
import sys, codecs, json
import openpyxl

# Setting up webdriver
driver = webdriver.Chrome("/Users/raunakbardia/Desktop/AirQuality/ExtractingData/Drivers/chromedriver")
url = "https://app.cpcbccr.com/ccr/#/caaqm-dashboard-all/caaqm-landing/caaqm-comparison-data"
driver.get(url)     # Opening CPCB Web Page
driver.maximize_window()

# Opening station list workbook
path = "Station_List.xlsx"
wb = openpyxl.load_workbook(path)
stationSheet = wb["StateCityPair"]   
nStations = len(stationSheet['A'])      # Number of rows in the state sheet to run nStates iterations

run = 0
startRow = 96
endRow = nStations + 1
for x in range(startRow,endRow):
    print("Reading Row Number ", x)
    stateName = stationSheet.cell(row = x, column = 1).value
    cityName =  stationSheet.cell(row = x, column = 2).value

    # Selecting the state
    if(run == 0):
        WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH, "/html/body/app-root/app-caaqm-dashboard/div/div/main/section/app-caaqm-comparison-data/div[1]/div/div[1]/div[1]/div/ng-select"))).click()
        run = 1
    else:
        driver.find_element_by_xpath("/html/body/app-root/app-caaqm-dashboard/div/div/main/section/app-caaqm-comparison-data/div[1]/div/div[1]/div[1]/div/ng-select").click()
    divElement = driver.find_element_by_class_name("filter")
    inputElement = divElement.find_element_by_xpath(".//input")
    inputElement.send_keys(stateName)
    inputElement.send_keys(Keys.RETURN)

    # Selecting the city
    driver.find_element_by_xpath("/html/body/app-root/app-caaqm-dashboard/div/div/main/section/app-caaqm-comparison-data/div[1]/div/div[1]/div[2]/div/ng-select").click()
    divElement = driver.find_element_by_class_name("filter")
    inputElement = divElement.find_element_by_xpath(".//input")
    inputElement.send_keys(cityName)
    if(x!=122):
        inputElement.send_keys(Keys.RETURN)
    else:   # Exception for Noida and Greater Noida
        driver.find_element_by_xpath("/html/body/app-root/app-caaqm-dashboard/div/div/main/section/app-caaqm-comparison-data/div[1]/div/div[1]/div[2]/div/ng-select/select-dropdown/div/div[2]/ul/li[2]").click()

    # Selecting all stations of the city
    driver.find_element_by_xpath("/html/body/app-root/app-caaqm-dashboard/div/div/main/section/app-caaqm-comparison-data/div[1]/div/div[2]/div[1]/div/div/multi-select/angular2-multiselect").click()
    driver.find_element_by_xpath("/html/body/app-root/app-caaqm-dashboard/div/div/main/section/app-caaqm-comparison-data/div[1]/div/div[2]/div[1]/div/div/multi-select/angular2-multiselect/div/div[2]/div[2]/div[1]/label/span[1]").click()
    
    # Selecting PM2.5, PM10, and Ozone parameters
    parameterDropdown = driver.find_element_by_xpath("/html/body/app-root/app-caaqm-dashboard/div/div/main/section/app-caaqm-comparison-data/div[1]/div/div[2]/div[2]/div/div/multi-select/angular2-multiselect")
    parameterDropdown.click()
    divElement = parameterDropdown.find_element_by_class_name("list-filter")
    inputElement = divElement.find_element_by_xpath(".//input")
    inputElement.send_keys("PM2.5")
    parameterDropdown.find_element_by_tag_name("li").click()
    inputElement.clear()
    inputElement.send_keys("PM10")
    parameterDropdown.find_element_by_tag_name("li").click()
    inputElement.clear()
    inputElement.send_keys("Ozone")
    parameterDropdown.find_element_by_tag_name("li").click()
    inputElement.clear()

    # Submitting the Query for Adding Station
    #this code will get the list of tags and select the second tag from the list
    element = driver.find_elements_by_tag_name('button')[1] 
    #this will click the element
    element.click()
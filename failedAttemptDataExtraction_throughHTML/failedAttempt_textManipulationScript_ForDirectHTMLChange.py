import requests
from bs4 import BeautifulSoup, Comment, NavigableString
import sys, codecs, json
import openpyxl

# Load the workbook
path = "Station_List.xlsx"
wb = openpyxl.load_workbook(path)
stationSheet = wb["StationList"]   
nStations = len(stationSheet['A'])      # Number of rows in the state sheet to run nStates iterations

f= open("queryText.txt","w+")
for x in range(2,nStations+1):
    stateName = stationSheet.cell(row = x, column = 1).value
    cityName =  stationSheet.cell(row = x, column = 2).value
    stationName = stationSheet.cell(row = x, column = 3).value
    str = "<tr _ngcontent-c3=\"\">\n\t \
        <td _ngcontent-c3=\"\">{0}</td>\n\t \
        <td _ngcontent-c3=\"\">{1}</td>\n\t \
        <td _ngcontent-c3=\"\">{2}</td>\n\t \
        <td _ngcontent-c3=\"\">PM2.5,PM10,Ozone</td>\n\t \
        <td _ngcontent-c3=\"\" align=\"center\"><i _ngcontent-c3=\"\" class=\"fa fa-times\" title=\"Remove\"></i></td>\n \
        </tr>\n"
    str = str.format(stateName, cityName, stationName)
    f.write(str)